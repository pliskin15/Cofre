"""
Conciliador PIX MAQUINETA — PMZ Peças e Pneus
Relaciona vendas PIX Maquineta (Cupom Fiscal, Nota Fiscal, Recibos)
com o extrato de Movimentação PIX Maquineta (XLSX Getnet/Santander).
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import pdfplumber
import os, re
from datetime import datetime
from theme import T, aplicar_estilos_ttk, aplicar_tags_tree, botao_tema, registrar_callback

# ─────────────────────────────────────────────────────────────────────────────
# CORES — obtidas dinamicamente via T() do theme.py
# ─────────────────────────────────────────────────────────────────────────────
def _cores():
    global COR_BG, COR_PAINEL, COR_BORDA, COR_ACENTO, COR_ACENTO2
    global COR_TEXTO, COR_TEXTO_SEC, COR_VERDE, COR_AMARELO
    global COR_VERMELHO, COR_CINZA, COR_AZUL, COR_LARANJA
    COR_BG        = T("BG")
    COR_PAINEL    = T("PAINEL")
    COR_BORDA     = T("BORDA")
    COR_ACENTO    = T("ACENTO")
    COR_ACENTO2   = T("ACENTO2")
    COR_TEXTO     = T("TEXTO")
    COR_TEXTO_SEC = T("TEXTO_SEC")
    COR_VERDE     = T("VERDE")
    COR_AMARELO   = T("AMARELO")
    COR_VERMELHO  = T("VERMELHO")
    COR_CINZA     = T("CINZA")
    COR_AZUL      = T("AZUL")
    COR_LARANJA   = T("LARANJA")

_cores()

# ─────────────────────────────────────────────────────────────────────────────
# PARSING DOS RELATÓRIOS
# ─────────────────────────────────────────────────────────────────────────────

def _num(val):
    """Converte string/float para float seguro. Suporta formato BR (1.234,56)."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if re.search(r"\d\.\d{3},\d", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except:
        return 0.0


def _extrair_linhas_pdf(path):
    linhas = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text(layout=True) or ""
            for linha in texto.splitlines():
                linhas.append(linha)
    return linhas


def _ultimo_num_linha(linha):
    tokens = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}|\d+\.\d{2}|\d+", linha)
    for tok in reversed(tokens):
        n = _num(tok)
        if n > 0:
            return n
    return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# PARSER: CUPOM FISCAL (PDF) — filtra VENDA PIX MAQUINETA
# ─────────────────────────────────────────────────────────────────────────────

def ler_cupom_fiscal(path):
    """
    Lê o Relatório Cupom Fiscal (PDF).
    Captura apenas linhas com "VENDA PIX" + "MAQUINETA" (exclui QR Code).
    """
    linhas = _extrair_linhas_pdf(path)
    registros = []

    cupom_atual  = None
    cond_atual   = ""
    vender_atual = ""

    for linha in linhas:
        linha_strip = linha.strip()
        if not linha_strip:
            continue
        up = linha_strip.upper()

        # ── Linha principal do cupom ──────────────────────────────────────────
        m = re.match(r"^(\d{5,})\s+\d{5,}\s+\d+\s+\w+\s+(\S+)\s+(\S+)", linha_strip)
        if m:
            cupom_atual  = m.group(1)
            cond_atual   = m.group(2)
            vender_atual = m.group(3)
            continue

        # ── Linhas de totais/rodapé → resetar contexto ────────────────────────
        if any(up.startswith(t) for t in ["TOTAL GERAL", "TOTAL :", "TOTAL:",
                                        "DESCRICAO", "TOTAIS", "CANCELADOS",
                                        "SERVICOS", "VENDAS"]):
            cupom_atual = None
            continue

        # ── Linha de pagamento VENDA PIX MAQUINETA ───────────────────────────
        # Ao contrário do QR Code, aqui exigimos "MAQUINETA" presente
        if "VENDA PIX" in up and "MAQUINETA" in up:
            if not cupom_atual:
                continue
            valor = _ultimo_num_linha(linha_strip)
            if valor > 0:
                registros.append({
                    "origem":     "Cupom Fiscal",
                    "referencia": f"Cupom {cupom_atual}",
                    "valor":      round(valor, 2),
                    "descricao":  (f"VENDA PIX MAQUINETA | Cupom {cupom_atual} "
                                   f"| Cond: {cond_atual} | Vend: {vender_atual}"),
                    "status":     "pendente",
                    "par_banco":  "",
                })
                cupom_atual = None

    return pd.DataFrame(registros)


# ─────────────────────────────────────────────────────────────────────────────
# PARSER: NOTA FISCAL (PDF) — filtra VENDA PIX MAQUINETA
# ─────────────────────────────────────────────────────────────────────────────

def ler_nota_fiscal(path):
    """
    Lê o Relatório de Venda Avista — Notas Fiscais (PDF).
    Captura apenas linhas com "VENDA PIX" + "MAQUINETA".
    """
    linhas = _extrair_linhas_pdf(path)
    registros = []

    nota_atual    = None
    cliente_atual = ""
    cond_atual    = ""

    for linha in linhas:
        linha_strip = linha.strip()
        if not linha_strip:
            continue
        up = linha_strip.upper()

        # ── Linha principal da nota ───────────────────────────────────────────
        m = re.match(r"^(\d{5,})\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(.+?)\s{2,}(\S+)\s+([\d\.,]+)", linha_strip)
        if m:
            nota_atual    = m.group(1)
            cliente_atual = m.group(6).strip()
            cond_atual    = m.group(7).strip()
            continue

        # Fallback
        if re.match(r"^\d{5,}\s", linha_strip):
            partes = linha_strip.split()
            nums_ini = sum(1 for p in partes[:5] if re.match(r"^\d+$", p))
            if nums_ini >= 3:
                nota_atual    = partes[0]
                cliente_atual = " ".join(p for p in partes[4:10]
                                         if not re.match(r"^[\d\.,]+$", p))
                cond_atual    = ""
            continue

        # ── Linha de pagamento VENDA PIX MAQUINETA ───────────────────────────
        if "VENDA PIX" in up and "MAQUINETA" in up:
            valor = _ultimo_num_linha(linha_strip)
            if valor > 0 and nota_atual:
                registros.append({
                    "origem":     "Nota Fiscal",
                    "referencia": f"NF {nota_atual}",
                    "valor":      round(valor, 2),
                    "descricao":  (f"VENDA PIX MAQUINETA | NF {nota_atual} "
                                   f"| {cliente_atual} | Cond: {cond_atual}"),
                    "status":     "pendente",
                    "par_banco":  "",
                })

    return pd.DataFrame(registros)


# ─────────────────────────────────────────────────────────────────────────────
# PARSER: RECIBOS (PDF) — filtra DEP. GETNET PIX
# ─────────────────────────────────────────────────────────────────────────────

def ler_recibos(path):
    """
    Lê o Relatório de Recibos (PDF).
    Captura apenas recibos com "DEP. GETNET PIX" (exclui DEP. PIX QRCOD).

    Layout de cada recibo:
      Linha cabeçalho:  41  41043849  3  41002201  NOME CLIENTE  DEP. GETNET PIX  TOTAL :  137,67
      Linha DPP:        DPP  2419196399  2  137,67  127,46  ...  DEPOSITO  ...

    Colunas DPP:
      [0]DPP [1]DOC [2]SERIE [3]RECEBIDO [4]V.DOC [5]JR.DOC [6]JR.CART
      [7]DESPESAS [8]DINH. [9]CHEQUE [10]CART.DEB [11]CART.CRED [12]DEPOSITO
      [13]ANTECIPADO [14]DEVCAR
    """
    linhas = _extrair_linhas_pdf(path)

    registros    = []
    recibo_atual = None

    for linha in linhas:
        linha_limpa = linha.strip()
        up = linha_limpa.upper()

        # ── Detecta início de recibo ──────────────────────────────────────────
        m_recibo = re.match(r"^(\d+)\s+(\d+)\s+", linha_limpa)
        tem_texto_apos = bool(re.search(r"[A-Za-z]", linha_limpa.split(None, 2)[-1])) \
                        if m_recibo else False
        if m_recibo and tem_texto_apos:

            # Salva recibo anterior se tiver valor
            if recibo_atual and recibo_atual["valor"] > 0:
                registros.append(recibo_atual)
            recibo_atual = None

            # Só captura recibos DEP. GETNET PIX
            if "DEP. GETNET PIX" in up:
                partes = linha_limpa.split()
                numero_recibo = partes[1]

                recibo_atual = {
                    "origem":     "Recibo",
                    "referencia": f"Recibo {numero_recibo}",
                    "valor":      0.0,
                    "descricao":  f"DEP. GETNET PIX | Recibo {numero_recibo}",
                    "status":     "pendente",
                    "par_banco":  "",
                }
            continue

        # ── Linhas DPP: soma coluna DEPOSITO (índice 9 dos valores com vírgula) ─
        if recibo_atual and (up.startswith("DPP") or up.startswith("ANT")):
            nums = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha_limpa)
            INDICE_DEPOSITO = 9
            if len(nums) > INDICE_DEPOSITO:
                deposito = _num(nums[INDICE_DEPOSITO])
                recibo_atual["valor"] = round(recibo_atual["valor"] + deposito, 2)

    # Adiciona último recibo
    if recibo_atual and recibo_atual["valor"] > 0:
        registros.append(recibo_atual)

    df = pd.DataFrame(registros)
    if not df.empty:
        df["saldo_rest"] = df["valor"]
    return df


# ─────────────────────────────────────────────────────────────────────────────
# PARSER: EXTRATO PIX MAQUINETA (XLSX — Getnet/Santander)
# ─────────────────────────────────────────────────────────────────────────────

def ler_mov_pix(path):
    """
    Lê o extrato de PIX Maquineta em formato XLSX (Getnet/Santander).

    Colunas relevantes:
      DATA/HORA DA VENDA | ID/TRANSAÇÃO (ID) | NÚMERO DO COMPROVANTE DE VENDAS (CV)
      NÚMERO DO TERMINAL | VALOR DA VENDA | STATUS

    Regras:
    - Cabeçalho está na linha de índice 7 do arquivo
    - Apenas registros com STATUS == "Paga" são importados
    - VALOR DA VENDA já é numérico (float) no xlsx
    """
    try:
        df_raw = pd.read_excel(path, sheet_name="PIX", header=7,
                               engine="openpyxl")
    except Exception:
        # Fallback: tenta sem especificar sheet
        df_raw = pd.read_excel(path, header=7, engine="openpyxl")

    # Normalizar nomes de colunas (remove espaços extras)
    df_raw.columns = [str(c).strip() for c in df_raw.columns]

    # Filtrar apenas transações "Paga"
    col_status = "STATUS"
    if col_status in df_raw.columns:
        df_raw = df_raw[df_raw[col_status].astype(str).str.strip().str.lower() == "paga"]

    # Remover linhas sem valor
    col_valor = "VALOR DA VENDA"
    df_raw = df_raw[pd.to_numeric(df_raw[col_valor], errors="coerce") > 0].copy()
    df_raw[col_valor] = pd.to_numeric(df_raw[col_valor], errors="coerce").round(2)

    # Montar DataFrame padronizado
    registros = []
    for _, row in df_raw.iterrows():
        dt_hora  = str(row.get("DATA/HORA DA VENDA", "")).strip()
        txid     = str(row.get("ID/TRANSAÇÃO (ID)", "")).strip()
        cv       = str(row.get("NÚMERO DO COMPROVANTE DE VENDAS (CV)", "")).strip()
        terminal = str(row.get("NÚMERO DO TERMINAL", "")).strip()
        valor    = float(row[col_valor])

        # Separar data e hora
        partes_dt = dt_hora.split(" ")
        dt   = partes_dt[0] if len(partes_dt) > 0 else ""
        hora = partes_dt[1] if len(partes_dt) > 1 else ""

        registros.append({
            "DT_RECEB":   dt,
            "HR_RECEB":   hora,
            "TERMINAL":   terminal,
            "CV":         cv,
            "TXID":       txid,
            "VALOR":      valor,
            "status":     "pendente",
            "par_venda":  "",
            "saldo_rest": valor,
        })

    COLS = ["DT_RECEB","HR_RECEB","TERMINAL","CV","TXID",
            "VALOR","status","par_venda","saldo_rest"]
    if not registros:
        df = pd.DataFrame(columns=COLS)
        df["VALOR"]      = pd.Series(dtype=float)
        df["saldo_rest"] = pd.Series(dtype=float)
    else:
        df = pd.DataFrame(registros, columns=COLS)
        df["saldo_rest"] = df["VALOR"]

    return df


# ─────────────────────────────────────────────────────────────────────────────
# CONCILIAÇÃO AUTOMÁTICA
# ─────────────────────────────────────────────────────────────────────────────

def conciliar_automatico(df_vendas, df_banco, tolerancia=0.01):
    dv = df_vendas.copy()
    db = df_banco.copy()

    dv["status"]    = "pendente"
    dv["par_banco"] = ""
    dv["saldo_rest"] = dv["valor"]

    db["status"]    = "pendente"
    db["par_venda"] = ""
    db["saldo_rest"] = db["VALOR"]

    par_counter = [0]

    def novo_par():
        par_counter[0] += 1
        return f"P{par_counter[0]:04d}"

    for iv, row_v in dv.iterrows():
        saldo_v = dv.at[iv, "saldo_rest"]
        if saldo_v <= tolerancia:
            continue

        candidatos = db[
            (db["saldo_rest"] > tolerancia) &
            (abs(db["VALOR"] - saldo_v) <= tolerancia)
        ]

        for ib, row_b in candidatos.iterrows():
            saldo_v = dv.at[iv, "saldo_rest"]
            saldo_b = db.at[ib, "saldo_rest"]
            if saldo_v <= tolerancia or saldo_b <= tolerancia:
                continue
            valor_match = min(saldo_v, saldo_b)
            par = novo_par()
            dv.at[iv, "par_banco"]  += ("," if dv.at[iv, "par_banco"] else "") + par
            dv.at[iv, "saldo_rest"]  = round(saldo_v - valor_match, 2)
            db.at[ib, "par_venda"]  += ("," if db.at[ib, "par_venda"] else "") + par
            db.at[ib, "saldo_rest"]  = round(saldo_b - valor_match, 2)

    def status_venda(row):
        if not row["par_banco"]:
            return "pendente"
        return "conciliado" if row["saldo_rest"] <= tolerancia else "parcial"

    def status_banco(row):
        if not row["par_venda"]:
            return "pendente"
        return "conciliado" if row["saldo_rest"] <= tolerancia else "parcial"

    dv["status"] = dv.apply(status_venda, axis=1)
    db["status"] = db.apply(status_banco, axis=1)

    return dv, db


# ─────────────────────────────────────────────────────────────────────────────
# INTERFACE
# ─────────────────────────────────────────────────────────────────────────────

class Conciliacao_PixMaquineta(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__()
        self.title("Conciliador PIX MAQUINETA — PMZ Peças e Pneus")
        self.geometry("1500x860")
        self.configure(bg=COR_BG)
        self.resizable(True, True)
        self.grab_set()   
        self.df_vendas = None
        self.df_banco  = None
        self.paths     = {"cupom": None, "nf": None, "recibo": None, "banco": None}

        self.sel_vendas = []
        self.sel_bancos = []

        self._build_ui()
        self._aplicar_estilos()

    # ─── Build UI ────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_topbar()
        corpo = tk.Frame(self, bg=COR_BG)
        corpo.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self._build_painel_esq(corpo)
        self._build_tabelas(corpo)
        self._build_statusbar()

    def _build_topbar(self):
        bar = tk.Frame(self, bg=COR_PAINEL, height=58)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        self._topbar = bar

        tk.Label(bar, text="🟠  Conciliador PIX MAQUINETA — PMZ",
                 bg=COR_PAINEL, fg=COR_TEXTO,
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=16, pady=12)

        btns = [
            ("🔄  Conciliar Auto",   COR_AZUL,    self.conciliar_auto),
            ("🤝  Conciliar Manual", COR_VERDE,   self.conciliar_manual),
            ("🔓  Desconciliar",     COR_AMARELO, self.desconciliar),
            ("🚫  Ignorar",          COR_CINZA,   self.ignorar),
        ]
        for txt, cor, cmd in btns:
            tk.Button(bar, text=txt, bg=cor, fg="white",
                      font=("Segoe UI", 9, "bold"), relief="flat",
                      padx=12, pady=6, cursor="hand2",
                      command=cmd).pack(side="left", padx=4, pady=12)

        # Botão exportar
        tk.Button(bar, text="📊  Exportar", bg=COR_LARANJA, fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.exportar).pack(side="left", padx=4, pady=12)

        # Botão de tema — alinhado à direita
        self._btn_tema = botao_tema(bar, callback=self._aplicar_tema)
        self._btn_tema.pack(side="right", padx=12, pady=12)

    def _build_painel_esq(self, parent):
        # Container externo (mantém a largura fixa)
        outer = tk.Frame(parent, bg=COR_PAINEL, width=230)
        outer.pack(side="left", fill="y", padx=(0, 10), pady=10)
        outer.pack_propagate(False)

        # Canvas + frame interno com scroll pela rodinha
        canvas = tk.Canvas(outer, bg=COR_PAINEL, highlightthickness=0,
                        width=230)
        canvas.pack(side="left", fill="both", expand=True)

        frame = tk.Frame(canvas, bg=COR_PAINEL)
        _win = canvas.create_window((0, 0), window=frame, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(_win, width=e.width)
        canvas.bind("<Configure>", _on_resize)

        def _on_frame_change(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        frame.bind("<Configure>", _on_frame_change)

        def _scroll(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _scroll)      # Windows / macOS
        canvas.bind_all("<Button-4>",                 # Linux scroll up
            lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>",                 # Linux scroll down
            lambda e: canvas.yview_scroll(1, "units"))

        tk.Label(frame, text="ARQUIVOS", bg=COR_PAINEL, fg=COR_ACENTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12, pady=(14, 4))

        self.lbl_paths = {}
        arquivos = [
            ("cupom",  "📄 Cupom Fiscal",    self.carregar_cupom),
            ("nf",     "🧾 Nota Fiscal",     self.carregar_nf),
            ("recibo", "📋 Recibos",         self.carregar_recibo),
            ("banco",  "🏦 PIX Maquineta",   self.carregar_banco),
        ]
        for chave, label, cmd in arquivos:
            tk.Button(frame, text=label, bg=COR_ACENTO2, fg="white",
                      font=("Segoe UI", 8, "bold"), relief="flat",
                      padx=8, pady=4, cursor="hand2", anchor="w",
                      command=cmd).pack(fill="x", padx=12, pady=(4, 0))
            lbl = tk.Label(frame, text="(não carregado)", bg=COR_PAINEL,
                           fg=COR_TEXTO_SEC, font=("Segoe UI", 7),
                           wraplength=200, justify="left")
            lbl.pack(anchor="w", padx=14, pady=(0, 4))
            self.lbl_paths[chave] = lbl

        # Resumo Vendas
        tk.Frame(frame, bg=COR_BORDA, height=1).pack(fill="x", padx=12, pady=10)
        tk.Label(frame, text="RESUMO VENDAS PIX MAQ.", bg=COR_PAINEL, fg=COR_ACENTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12)

        self.lbl_res = {}
        for k, lbl, cor in [
            ("v_total",      "Total registros:",  COR_TEXTO),
            ("v_conciliado", "✅ Conciliados:",   COR_VERDE),
            ("v_parcial",    "⚠ Parciais:",       COR_AMARELO),
            ("v_pendente",   "❌ Pendentes:",     COR_VERMELHO),
            ("v_ignorado",   "🚫 Ignorados:",     COR_CINZA),
            ("v_soma",       "Σ Valor Vendas:",   COR_TEXTO),
        ]:
            row = tk.Frame(frame, bg=COR_PAINEL)
            row.pack(fill="x", padx=12, pady=1)
            tk.Label(row, text=lbl, bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                     font=("Segoe UI", 8)).pack(side="left")
            l = tk.Label(row, text="—", bg=COR_PAINEL, fg=cor,
                         font=("Segoe UI", 8, "bold"))
            l.pack(side="right")
            self.lbl_res[k] = l

        # Resumo Banco
        tk.Frame(frame, bg=COR_BORDA, height=1).pack(fill="x", padx=12, pady=6)
        tk.Label(frame, text="RESUMO PIX MAQUINETA", bg=COR_PAINEL, fg=COR_ACENTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12)

        for k, lbl, cor in [
            ("b_total",      "Total registros:",  COR_TEXTO),
            ("b_conciliado", "✅ Conciliados:",   COR_VERDE),
            ("b_pendente",   "❌ Pendentes:",     COR_VERMELHO),
            ("b_soma",       "Σ Valor Banco:",    COR_TEXTO),
            ("diferenca",    "Δ Diferença:",      COR_AZUL),
        ]:
            row = tk.Frame(frame, bg=COR_PAINEL)
            row.pack(fill="x", padx=12, pady=1)
            tk.Label(row, text=lbl, bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                     font=("Segoe UI", 8)).pack(side="left")
            l = tk.Label(row, text="—", bg=COR_PAINEL, fg=cor,
                         font=("Segoe UI", 8, "bold"))
            l.pack(side="right")
            self.lbl_res[k] = l

        # Conciliação manual
        tk.Frame(frame, bg=COR_BORDA, height=1).pack(fill="x", padx=12, pady=8)
        tk.Label(frame, text="MANUAL", bg=COR_PAINEL, fg=COR_ACENTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12)
        tk.Label(frame,
                 text="1. Clique em uma ou mais VENDAS (tabela cima)\n"
                      "   (clique novamente para desmarcar)\n"
                      "2. Clique em um ou mais PIX (tabela baixo)\n"
                      "3. Pressione 'Conciliar Manual'",
                 bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8), justify="left").pack(anchor="w", padx=12, pady=2)

        self.lbl_sel_v = tk.Label(frame, text="Venda: (nenhuma)",
                                  bg=COR_PAINEL, fg=COR_VERDE,
                                  font=("Segoe UI", 8, "italic"), wraplength=210)
        self.lbl_sel_v.pack(anchor="w", padx=12)
        self.lbl_sel_b = tk.Label(frame, text="PIX banco: (nenhum)",
                                  bg=COR_PAINEL, fg=COR_AZUL,
                                  font=("Segoe UI", 8, "italic"), wraplength=210)
        self.lbl_sel_b.pack(anchor="w", padx=12, pady=(2, 0))

        tk.Button(frame, text="Limpar seleção", bg=COR_BG, fg=COR_TEXTO_SEC,
                  font=("Segoe UI", 8), relief="flat", cursor="hand2",
                  command=self.limpar_selecao).pack(anchor="w", padx=12, pady=(6, 0))

        # ── Filtros VENDAS ──────────────────────────────────────────────────
        tk.Frame(frame, bg=COR_BORDA, height=1).pack(fill="x", padx=12, pady=8)
        tk.Label(frame, text="FILTROS — VENDAS", bg=COR_PAINEL, fg=COR_ACENTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12)

        tk.Label(frame, text="Status:", bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=12, pady=(4, 0))
        self.filtro_status_v = ttk.Combobox(frame, state="readonly",
            values=["Todos", "pendente", "parcial", "conciliado", "ignorado"])
        self.filtro_status_v.set("Todos")
        self.filtro_status_v.pack(fill="x", padx=12, pady=(2, 4))
        self.filtro_status_v.bind("<<ComboboxSelected>>", lambda _: self.atualizar_tabelas())

        tk.Label(frame, text="Valor (R$):", bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=12)
        self.filtro_valor_v = tk.Entry(frame, bg=COR_BG, fg=COR_TEXTO,
                                       font=("Segoe UI", 8), relief="flat",
                                       insertbackground=COR_TEXTO)
        self.filtro_valor_v.pack(fill="x", padx=12, pady=(2, 0))
        self.filtro_valor_v.bind("<KeyRelease>", lambda _: self.atualizar_tabelas())
        tk.Label(frame, text="(ex: 150.00 ou deixe vazio)", bg=COR_PAINEL,
                 fg=COR_TEXTO_SEC, font=("Segoe UI", 7)).pack(anchor="w", padx=12)

        # ── Filtros PAGAMENTOS (Banco) ───────────────────────────────────────
        tk.Frame(frame, bg=COR_BORDA, height=1).pack(fill="x", padx=12, pady=8)
        tk.Label(frame, text="FILTROS — PIX MAQUINETA", bg=COR_PAINEL, fg=COR_LARANJA,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12)

        tk.Label(frame, text="Data (DT_RECEB):", bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=12, pady=(4, 0))
        self.filtro_data_b = tk.Entry(frame, bg=COR_BG, fg=COR_TEXTO,
                                      font=("Segoe UI", 8), relief="flat",
                                      insertbackground=COR_TEXTO)
        self.filtro_data_b.pack(fill="x", padx=12, pady=(2, 0))
        self.filtro_data_b.bind("<KeyRelease>", lambda _: self.atualizar_tabelas())
        tk.Label(frame, text="(ex: 2024-05-10 ou 10/05)", bg=COR_PAINEL,
                 fg=COR_TEXTO_SEC, font=("Segoe UI", 7)).pack(anchor="w", padx=12)

        tk.Label(frame, text="Status:", bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=12, pady=(4, 0))
        self.filtro_status_b = ttk.Combobox(frame, state="readonly",
            values=["Todos", "pendente", "parcial", "conciliado", "ignorado"])
        self.filtro_status_b.set("Todos")
        self.filtro_status_b.pack(fill="x", padx=12, pady=(2, 4))
        self.filtro_status_b.bind("<<ComboboxSelected>>", lambda _: self.atualizar_tabelas())

        tk.Label(frame, text="Valor (R$):", bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=12)
        self.filtro_valor_b = tk.Entry(frame, bg=COR_BG, fg=COR_TEXTO,
                                       font=("Segoe UI", 8), relief="flat",
                                       insertbackground=COR_TEXTO)
        self.filtro_valor_b.pack(fill="x", padx=12, pady=(2, 0))
        self.filtro_valor_b.bind("<KeyRelease>", lambda _: self.atualizar_tabelas())
        tk.Label(frame, text="(ex: 250.00 ou deixe vazio)", bg=COR_PAINEL,
                 fg=COR_TEXTO_SEC, font=("Segoe UI", 7)).pack(anchor="w", padx=12)


    def _build_tabelas(self, parent):
        frame = tk.Frame(parent, bg=COR_BG)
        frame.pack(side="left", fill="both", expand=True, pady=10)

        # ── Tabela superior: Vendas ───────────────────────────────────────────
        tk.Label(frame, text="VENDAS PIX MAQUINETA  (Cupom Fiscal + Nota Fiscal + Recibos)",
                 bg=COR_BG, fg=COR_ACENTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 2))

        frm_v = tk.Frame(frame, bg=COR_BG)
        frm_v.pack(fill="both", expand=True)

        self.cols_v = ["origem", "referencia", "valor", "saldo_rest",
                       "descricao", "status", "par_banco"]
        self.tree_v = ttk.Treeview(frm_v, columns=self.cols_v, show="headings",
                                   selectmode="extended", height=12)
        largs_v = {"origem": 80, "referencia": 100, "valor": 90, "saldo_rest": 90,
                   "descricao": 380, "status": 90, "par_banco": 90}
        for col in self.cols_v:
            self.tree_v.heading(col, text=col.upper())
            self.tree_v.column(col, width=largs_v.get(col, 100),
                               anchor="center" if col in ("valor", "saldo_rest",
                                                          "status", "par_banco") else "w")

        sb_vy = ttk.Scrollbar(frm_v, orient="vertical",   command=self.tree_v.yview)
        sb_vx = ttk.Scrollbar(frm_v, orient="horizontal", command=self.tree_v.xview)
        self.tree_v.configure(yscrollcommand=sb_vy.set, xscrollcommand=sb_vx.set)
        sb_vy.pack(side="right",  fill="y")
        sb_vx.pack(side="bottom", fill="x")
        self.tree_v.pack(fill="both", expand=True)
        self.tree_v.bind("<ButtonRelease-1>", self._on_click_venda)
        self._cfg_tags(self.tree_v)

        # ── Tabela inferior: PIX Maquineta ────────────────────────────────────
        tk.Label(frame, text="EXTRATO PIX MAQUINETA — BANCO (XLSX)",
                 bg=COR_BG, fg=COR_LARANJA,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(8, 2))

        frm_b = tk.Frame(frame, bg=COR_BG)
        frm_b.pack(fill="both", expand=True)

        self.cols_b = ["DT_RECEB", "HR_RECEB", "TERMINAL", "CV",
                       "TXID", "VALOR", "saldo_rest", "status", "par_venda"]
        self.tree_b = ttk.Treeview(frm_b, columns=self.cols_b, show="headings",
                                   selectmode="extended", height=12)
        largs_b = {"DT_RECEB": 90, "HR_RECEB": 60, "TERMINAL": 90, "CV": 90,
                   "TXID": 260, "VALOR": 90, "saldo_rest": 90,
                   "status": 90, "par_venda": 90}
        for col in self.cols_b:
            self.tree_b.heading(col, text=col.upper())
            self.tree_b.column(col, width=largs_b.get(col, 90),
                               anchor="center" if col in ("VALOR", "saldo_rest",
                                                          "status", "par_venda",
                                                          "HR_RECEB", "DT_RECEB") else "w")

        sb_by = ttk.Scrollbar(frm_b, orient="vertical",   command=self.tree_b.yview)
        sb_bx = ttk.Scrollbar(frm_b, orient="horizontal", command=self.tree_b.xview)
        self.tree_b.configure(yscrollcommand=sb_by.set, xscrollcommand=sb_bx.set)
        sb_by.pack(side="right",  fill="y")
        sb_bx.pack(side="bottom", fill="x")
        self.tree_b.pack(fill="both", expand=True)
        self.tree_b.bind("<ButtonRelease-1>", self._on_click_banco)
        self._cfg_tags(self.tree_b)

    def _cfg_tags(self, tree):
        aplicar_tags_tree(tree)

    def _build_statusbar(self):
        bar = tk.Frame(self, bg=COR_PAINEL, height=26)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self.status_var = tk.StringVar(value="Pronto. Carregue os relatórios para começar.")
        tk.Label(bar, textvariable=self.status_var, bg=COR_PAINEL, fg=COR_TEXTO_SEC,
                 font=("Segoe UI", 8), anchor="w").pack(side="left", padx=10)

    def _aplicar_estilos(self):
        aplicar_estilos_ttk(ttk.Style(self))
        registrar_callback(self._aplicar_tema)

    def _aplicar_tema(self):
        """Reaplicar cores de todos os widgets após troca de tema."""
        _cores()
        self.configure(bg=T("BG"))
        aplicar_estilos_ttk(ttk.Style(self))
        for tree in (self.tree_v, self.tree_b):
            aplicar_tags_tree(tree)
        from theme import recolorir_widget
        recolorir_widget(self)
        self.atualizar_tabelas()

    # ─── Carregamento de arquivos ─────────────────────────────────────────────

    def _carregar(self, chave, func, label, eh_pdf=False):
        ft_pdf   = [("PDF",   "*.pdf"),              ("Todos", "*.*")]
        ft_excel = [("Excel", "*.xlsx *.xls *.xlsm"), ("Todos", "*.*")]
        path = filedialog.askopenfilename(
            title=f"Selecionar {label}",
            filetypes=ft_pdf if eh_pdf else ft_excel)
        if not path:
            return
        try:
            self.status_var.set(f"⏳ Carregando {label}...")
            self.update()
            df = func(path)
            self.paths[chave] = path
            n = len(df)
            self.lbl_paths[chave].config(
                text=f"✅ {os.path.basename(path)} ({n} reg.)")
            self.status_var.set(
                f"✅ {label} carregado — {n} registros encontrados.")
            return df
        except Exception as e:
            import traceback
            messagebox.showerror("Erro",
                f"Erro ao carregar {label}:\n{e}\n\n{traceback.format_exc()}")
            return None

    def _unificar_vendas(self):
        partes = []
        for chave in ("cupom", "nf", "recibo"):
            df = getattr(self, f"_df_{chave}", None)
            if df is not None:
                partes.append(df)
        if not partes:
            return None
        df = pd.concat(partes, ignore_index=True)
        df["status"]     = "pendente"
        df["par_banco"]  = ""
        df["saldo_rest"] = df["valor"]
        return df

    def carregar_cupom(self):
        df = self._carregar("cupom", ler_cupom_fiscal, "Cupom Fiscal", eh_pdf=True)
        if df is not None:
            self._df_cupom = df
            self._recarregar_vendas()

    def carregar_nf(self):
        df = self._carregar("nf", ler_nota_fiscal, "Nota Fiscal", eh_pdf=True)
        if df is not None:
            self._df_nf = df
            self._recarregar_vendas()

    def carregar_recibo(self):
        df = self._carregar("recibo", ler_recibos, "Recibos", eh_pdf=True)
        if df is not None:
            self._df_recibo = df
            self._recarregar_vendas()

    def carregar_banco(self):
        # Extrato maquineta é XLSX (não PDF)
        df = self._carregar("banco", ler_mov_pix, "PIX Maquineta (XLSX)", eh_pdf=False)
        if df is not None:
            self.df_banco = df
            self.atualizar_tabelas()
            self.atualizar_resumo()

    def _recarregar_vendas(self):
        self.df_vendas = self._unificar_vendas()
        self.limpar_selecao()
        self.atualizar_tabelas()
        self.atualizar_resumo()

    # ─── Ações ───────────────────────────────────────────────────────────────

    def conciliar_auto(self):
        if self.df_vendas is None or self.df_banco is None:
            messagebox.showwarning("Aviso",
                "Carregue os relatórios de vendas e o extrato do banco primeiro.")
            return
        self.status_var.set("⏳ Conciliando automaticamente...")
        self.update()
        self.df_vendas, self.df_banco = conciliar_automatico(self.df_vendas, self.df_banco)
        self.limpar_selecao()
        self.atualizar_tabelas()
        self.atualizar_resumo()
        n_cv = (self.df_vendas["status"] == "conciliado").sum()
        n_pv = (self.df_vendas["status"] == "pendente").sum()
        n_cb = (self.df_banco["status"]  == "conciliado").sum()
        n_pb = (self.df_banco["status"]  == "pendente").sum()
        self.status_var.set(
            f"🔄 Auto concluída — Vendas: ✅{n_cv} ❌{n_pv} | Banco: ✅{n_cb} ❌{n_pb}")

    def conciliar_manual(self):
        if not self.sel_vendas or not self.sel_bancos:
            messagebox.showwarning("Seleção incompleta",
                "Selecione pelo menos uma VENDA e pelo menos um registro PIX do banco.")
            return
        tol = 0.01
        for iv in self.sel_vendas:
            for ib in self.sel_bancos:
                sv = self.df_vendas.at[iv, "saldo_rest"]
                sb = self.df_banco.at[ib, "saldo_rest"]
                if sv <= tol or sb <= tol:
                    continue
                val = min(sv, sb)
                par = f"M{iv}-{ib}"
                def ap(df, idx, col, p):
                    a = df.at[idx, col]
                    df.at[idx, col] = f"{a},{p}" if a else p
                ap(self.df_vendas, iv, "par_banco", par)
                ap(self.df_banco,  ib, "par_venda", par)
                self.df_vendas.at[iv, "saldo_rest"] = round(sv - val, 2)
                self.df_banco.at[ib, "saldo_rest"]  = round(sb - val, 2)

        def novo_st_v(idx):
            r = self.df_vendas.loc[idx]
            if not r["par_banco"]: return "pendente"
            return "conciliado" if r["saldo_rest"] <= tol else "parcial"

        def novo_st_b(idx):
            r = self.df_banco.loc[idx]
            if not r["par_venda"]: return "pendente"
            return "conciliado" if r["saldo_rest"] <= tol else "parcial"

        for iv in self.sel_vendas:
            self.df_vendas.at[iv, "status"] = novo_st_v(iv)
        for ib in self.sel_bancos:
            self.df_banco.at[ib, "status"] = novo_st_b(ib)

        self.limpar_selecao()
        self.atualizar_tabelas()
        self.atualizar_resumo()
        self.status_var.set("🤝 Conciliação manual aplicada.")

    def desconciliar(self):
        self._alterar_status_selecionados(None)

    def ignorar(self):
        self._alterar_status_selecionados("ignorado")

    def _alterar_status_selecionados(self, novo_status):
        for tree, df, col_par in [
            (self.tree_v, self.df_vendas, "par_banco"),
            (self.tree_b, self.df_banco,  "par_venda"),
        ]:
            if df is None:
                continue
            for item in tree.selection():
                idx = self._item_id(tree, item)
                if idx is None:
                    continue
                if novo_status is None:
                    orig = df.at[idx, "valor"] if "valor" in df.columns else df.at[idx, "VALOR"]
                    df.at[idx, "status"]     = "pendente"
                    df.at[idx, col_par]      = ""
                    df.at[idx, "saldo_rest"] = orig
                else:
                    df.at[idx, "status"] = novo_status
        self.atualizar_tabelas()
        self.atualizar_resumo()
        acao = "desconciliados" if novo_status is None else "ignorados"
        self.status_var.set(f"✔ Registros {acao}.")

    # ─── Cliques na tabela ────────────────────────────────────────────────────

    def _on_click_venda(self, event):
        item = self.tree_v.identify_row(event.y)
        if not item:
            return
        idx = self._item_id(self.tree_v, item)
        if idx is None:
            return
        if idx in self.sel_vendas:
            self.sel_vendas.remove(idx)
        else:
            self.sel_vendas.append(idx)
        n = len(self.sel_vendas)
        if n == 0:
            self.lbl_sel_v.config(text="Venda: (nenhuma)")
        elif n == 1:
            row = self.df_vendas.loc[self.sel_vendas[0]]
            self.lbl_sel_v.config(
                text=f"Venda: R$ {row['valor']:,.2f} | {str(row['referencia'])[:30]}")
        else:
            soma = sum(self.df_vendas.at[i, "valor"] for i in self.sel_vendas)
            self.lbl_sel_v.config(
                text=f"Vendas: {n} selecionada(s) | Σ R$ {soma:,.2f}")
        self._destacar()

    def _on_click_banco(self, event):
        item = self.tree_b.identify_row(event.y)
        if not item:
            return
        idx = self._item_id(self.tree_b, item)
        if idx is None:
            return
        if self.sel_vendas:
            if idx in self.sel_bancos:
                self.sel_bancos.remove(idx)
            else:
                self.sel_bancos.append(idx)
            self.lbl_sel_b.config(
                text=f"PIX banco: {len(self.sel_bancos)} selecionado(s)")
        self._destacar()

    def _destacar(self):
        if self.df_vendas is not None:
            for item in self.tree_v.get_children():
                idx = self._item_id(self.tree_v, item)
                if idx is None: continue
                tag = self.df_vendas.at[idx, "status"]
                self.tree_v.item(item, tags=(tag,))
            for iv in self.sel_vendas:
                it = self._buscar_item(self.tree_v, iv)
                if it:
                    self.tree_v.item(it, tags=("selecionado",))
        if self.df_banco is not None:
            for item in self.tree_b.get_children():
                idx = self._item_id(self.tree_b, item)
                if idx is None: continue
                tag = self.df_banco.at[idx, "status"]
                self.tree_b.item(item, tags=(tag,))
            for ib in self.sel_bancos:
                it = self._buscar_item(self.tree_b, ib)
                if it:
                    self.tree_b.item(it, tags=("selecionado",))

    def limpar_selecao(self):
        self.sel_vendas = []
        self.sel_bancos = []
        self.lbl_sel_v.config(text="Venda: (nenhuma)")
        self.lbl_sel_b.config(text="PIX banco: (nenhum)")

    # ─── Atualização das tabelas ──────────────────────────────────────────────

    def atualizar_tabelas(self):
        self._popular_tree_vendas()
        self._popular_tree_banco()

    def _popular_tree_vendas(self):
        self._map_v  = {}
        self._rmap_v = {}
        self.tree_v.delete(*self.tree_v.get_children())
        if self.df_vendas is None:
            return
        df = self._filtrar(self.df_vendas, "status", tabela="vendas")
        for _, (idx, row) in enumerate(df.iterrows()):
            vals = (
                row.get("origem",     ""),
                row.get("referencia", ""),
                f"{row['valor']:,.2f}",
                f"{row['saldo_rest']:,.2f}",
                str(row.get("descricao", ""))[:70],
                row.get("status",    ""),
                row.get("par_banco", ""),
            )
            st  = row.get("status", "pendente")
            tag = st if st in ("conciliado", "parcial", "pendente", "ignorado") else "pendente"
            item = self.tree_v.insert("", "end", values=vals, tags=(tag,))
            self._map_v[item]  = idx
            self._rmap_v[idx]  = item

    def _popular_tree_banco(self):
        self._map_b  = {}
        self._rmap_b = {}
        self.tree_b.delete(*self.tree_b.get_children())
        if self.df_banco is None:
            return
        df = self._filtrar(self.df_banco, "status", tabela="banco")
        for _, (idx, row) in enumerate(df.iterrows()):
            vals = (
                str(row.get("DT_RECEB",  "")),
                str(row.get("HR_RECEB",  "")),
                str(row.get("TERMINAL",  "")),
                str(row.get("CV",        "")),
                str(row.get("TXID",      ""))[:55],
                f"{row['VALOR']:,.2f}",
                f"{row['saldo_rest']:,.2f}",
                row.get("status",    ""),
                row.get("par_venda", ""),
            )
            st  = row.get("status", "pendente")
            tag = st if st in ("conciliado", "parcial", "pendente", "ignorado") else "pendente"
            item = self.tree_b.insert("", "end", values=vals, tags=(tag,))
            self._map_b[item]  = idx
            self._rmap_b[idx]  = item

    def _filtrar(self, df, col_status, tabela="vendas"):
        result = df.copy()

        if tabela == "vendas":
            f_status = self.filtro_status_v.get()
            f_valor  = self.filtro_valor_v.get().strip()
        else:
            f_status = self.filtro_status_b.get()
            f_valor  = self.filtro_valor_b.get().strip()
            f_data   = self.filtro_data_b.get().strip()
            if f_data:
                result = result[result["DT_RECEB"].astype(str).str.contains(f_data, na=False)]

        if f_status != "Todos":
            result = result[result[col_status] == f_status]

        if f_valor:
            try:
                v = float(f_valor.replace(",", "."))
                col_v = "valor" if tabela == "vendas" else "VALOR"
                result = result[abs(result[col_v] - v) < 0.01]
            except ValueError:
                pass

        return result


    def atualizar_resumo(self):
        if self.df_vendas is not None:
            cnt = self.df_vendas["status"].value_counts()
            self.lbl_res["v_total"].config(text=str(len(self.df_vendas)))
            for s in ("conciliado", "parcial", "pendente", "ignorado"):
                self.lbl_res[f"v_{s}"].config(text=str(cnt.get(s, 0)))
            self.lbl_res["v_soma"].config(
                text=f"R$ {self.df_vendas['valor'].sum():,.2f}")

        if self.df_banco is not None:
            cnt = self.df_banco["status"].value_counts()
            self.lbl_res["b_total"].config(text=str(len(self.df_banco)))
            self.lbl_res["b_conciliado"].config(text=str(cnt.get("conciliado", 0)))
            self.lbl_res["b_pendente"].config(text=str(cnt.get("pendente", 0)))
            soma_b = self.df_banco["VALOR"].sum()
            self.lbl_res["b_soma"].config(text=f"R$ {soma_b:,.2f}")

            if self.df_vendas is not None:
                soma_v = self.df_vendas["valor"].sum()
                dif    = soma_v - soma_b
                cor    = COR_VERDE if abs(dif) < 0.05 else COR_VERMELHO
                self.lbl_res["diferenca"].config(
                    text=f"R$ {dif:,.2f}", fg=cor)

    # ─── Exportação ──────────────────────────────────────────────────────────

    def exportar(self):
        """Abre janela para informar o número da loja e exporta resumo em XLSX."""
        if self.df_vendas is None and self.df_banco is None:
            messagebox.showwarning("Aviso", "Carregue pelo menos um relatório antes de exportar.")
            return

        # ── Janela para número da loja ────────────────────────────────────────
        dlg = tk.Toplevel(self)
        dlg.title("Exportar Relatório")
        dlg.geometry("340x160")
        dlg.resizable(False, False)
        dlg.configure(bg=COR_PAINEL)
        dlg.grab_set()
        dlg.transient(self)

        tk.Label(dlg, text="Número da Loja:", bg=COR_PAINEL, fg=COR_TEXTO,
                 font=("Segoe UI", 10, "bold")).pack(pady=(20, 4))
        entry_loja = tk.Entry(dlg, font=("Segoe UI", 11), width=14,
                               bg=COR_BG, fg=COR_TEXTO, insertbackground=COR_TEXTO,
                               relief="flat", justify="center")
        entry_loja.pack(pady=4)
        entry_loja.focus_set()

        resultado = {"loja": None}

        def confirmar(event=None):
            v = entry_loja.get().strip()
            if not v:
                messagebox.showwarning("Aviso", "Informe o número da loja.", parent=dlg)
                return
            resultado["loja"] = v
            dlg.destroy()

        entry_loja.bind("<Return>", confirmar)
        tk.Button(dlg, text="Exportar", bg=COR_ACENTO, fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat",
                  padx=16, pady=6, cursor="hand2",
                  command=confirmar).pack(pady=10)

        dlg.wait_window()
        if not resultado["loja"]:
            return

        num_loja = resultado["loja"]

        # ── Escolher destino ──────────────────────────────────────────────────
        data_hoje = datetime.now().strftime("%Y%m%d")
        nome_sugerido = f"ConciliacaoPIX_Loja{num_loja}_{data_hoje}.xlsx"
        path_out = filedialog.asksaveasfilename(
            title="Salvar relatório",
            defaultextension=".xlsx",
            initialfile=nome_sugerido,
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if not path_out:
            return

        # ── Montar dados ──────────────────────────────────────────────────────
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # ── Aba 1: Resumo Geral ───────────────────────────────────────────
            ws_res = wb.active
            ws_res.title = "Resumo"

            def hdr_style(cell, cor_hex="2B4590"):
                cell.font      = Font(bold=True, color="FFFFFF", size=10)
                cell.fill      = PatternFill("solid", fgColor=cor_hex)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            def sub_hdr(cell):
                cell.font      = Font(bold=True, size=9)
                cell.fill      = PatternFill("solid", fgColor="D9E1F2")
                cell.alignment = Alignment(horizontal="left")

            def val_cell(cell, bold=False):
                cell.alignment = Alignment(horizontal="right")
                if bold:
                    cell.font = Font(bold=True)

            thin = Side(style="thin", color="CCCCCC")
            borda = Border(left=thin, right=thin, top=thin, bottom=thin)

            now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

            # Cabeçalho do documento
            ws_res["A1"] = f"Conciliação PIX Maquineta — Loja {num_loja}"
            ws_res["A1"].font = Font(bold=True, size=13, color="2B4590")
            ws_res["A2"] = f"Gerado em: {now_str}"
            ws_res["A2"].font = Font(size=9, color="888888")
            ws_res.merge_cells("A1:D1")
            ws_res.merge_cells("A2:D2")

            row = 4

            # ── Bloco Vendas ──────────────────────────────────────────────────
            ws_res.cell(row, 1, "VENDAS PIX MAQUINETA")
            hdr_style(ws_res.cell(row, 1))
            ws_res.merge_cells(f"A{row}:B{row}")
            row += 1

            if self.df_vendas is not None:
                dv = self.df_vendas
                cnt_v = dv["status"].value_counts()
                soma_v = dv["valor"].sum()
                soma_conc_v = dv.loc[dv["status"] == "conciliado", "valor"].sum()
                soma_pend_v = dv.loc[dv["status"] == "pendente",   "valor"].sum()
                soma_parc_v = dv.loc[dv["status"] == "parcial",    "valor"].sum()
                soma_ign_v  = dv.loc[dv["status"] == "ignorado",   "valor"].sum()

                items_v = [
                    ("Total de registros",       len(dv),                         False),
                    ("✅ Conciliados",            cnt_v.get("conciliado", 0),      False),
                    ("⚠ Parciais",               cnt_v.get("parcial", 0),         False),
                    ("❌ Pendentes",              cnt_v.get("pendente", 0),        False),
                    ("🚫 Ignorados",              cnt_v.get("ignorado", 0),        False),
                    ("Σ Valor Total Vendas",      f"R$ {soma_v:,.2f}",            True),
                    ("Σ Valor Conciliado",        f"R$ {soma_conc_v:,.2f}",       False),
                    ("Σ Valor Pendente",          f"R$ {soma_pend_v:,.2f}",       False),
                    ("Σ Valor Parcial",           f"R$ {soma_parc_v:,.2f}",       False),
                    ("Σ Valor Ignorado",          f"R$ {soma_ign_v:,.2f}",        False),
                    ("% Conciliado",              f"{100*soma_conc_v/soma_v:.1f}%" if soma_v else "—", False),
                ]
            else:
                items_v = [("(sem dados)", "—", False)]

            for label, valor, bold in items_v:
                c1 = ws_res.cell(row, 1, label)
                c2 = ws_res.cell(row, 2, valor)
                sub_hdr(c1)
                val_cell(c2, bold)
                c1.border = borda
                c2.border = borda
                row += 1

            row += 1

            # ── Bloco Banco ───────────────────────────────────────────────────
            ws_res.cell(row, 1, "EXTRATO PIX MAQUINETA — BANCO")
            hdr_style(ws_res.cell(row, 1), cor_hex="B45309")
            ws_res.merge_cells(f"A{row}:B{row}")
            row += 1

            if self.df_banco is not None:
                db = self.df_banco
                cnt_b = db["status"].value_counts()
                soma_b = db["VALOR"].sum()
                soma_conc_b = db.loc[db["status"] == "conciliado", "VALOR"].sum()
                soma_pend_b = db.loc[db["status"] == "pendente",   "VALOR"].sum()

                items_b = [
                    ("Total de registros",       len(db),                         False),
                    ("✅ Conciliados",            cnt_b.get("conciliado", 0),      False),
                    ("❌ Pendentes",              cnt_b.get("pendente", 0),        False),
                    ("Σ Valor Total Banco",       f"R$ {soma_b:,.2f}",            True),
                    ("Σ Valor Conciliado",        f"R$ {soma_conc_b:,.2f}",       False),
                    ("Σ Valor Pendente",          f"R$ {soma_pend_b:,.2f}",       False),
                    ("% Conciliado",              f"{100*soma_conc_b/soma_b:.1f}%" if soma_b else "—", False),
                ]
            else:
                items_b = [("(sem dados)", "—", False)]

            for label, valor, bold in items_b:
                c1 = ws_res.cell(row, 1, label)
                c2 = ws_res.cell(row, 2, valor)
                sub_hdr(c1)
                val_cell(c2, bold)
                c1.border = borda
                c2.border = borda
                row += 1

            row += 1

            # ── Bloco Diferença ───────────────────────────────────────────────
            if self.df_vendas is not None and self.df_banco is not None:
                ws_res.cell(row, 1, "DIFERENÇA")
                hdr_style(ws_res.cell(row, 1), cor_hex="1D6A3A")
                ws_res.merge_cells(f"A{row}:B{row}")
                row += 1

                dif = soma_v - soma_b
                cor_dif = "1D6A3A" if abs(dif) < 0.05 else "C0392B"
                c1 = ws_res.cell(row, 1, "Δ Vendas − Banco")
                c2 = ws_res.cell(row, 2, f"R$ {dif:,.2f}")
                sub_hdr(c1)
                c2.font      = Font(bold=True, color=cor_dif)
                c2.alignment = Alignment(horizontal="right")
                c1.border = borda
                c2.border = borda

            ws_res.column_dimensions["A"].width = 32
            ws_res.column_dimensions["B"].width = 22

            # ── Aba 2: Vendas detalhado ───────────────────────────────────────
            if self.df_vendas is not None:
                ws_v = wb.create_sheet("Vendas")
                cols_v_exp = ["origem", "referencia", "valor", "saldo_rest",
                               "descricao", "status", "par_banco"]
                for ci, col in enumerate(cols_v_exp, 1):
                    c = ws_v.cell(1, ci, col.upper())
                    hdr_style(c)

                for ri, (_, row_d) in enumerate(self.df_vendas.iterrows(), 2):
                    for ci, col in enumerate(cols_v_exp, 1):
                        v = row_d.get(col, "")
                        ws_v.cell(ri, ci, v)

                for ci, col in enumerate(cols_v_exp, 1):
                    ws_v.column_dimensions[get_column_letter(ci)].width = (
                        50 if col == "descricao" else 16)

            # ── Aba 3: Banco detalhado ────────────────────────────────────────
            if self.df_banco is not None:
                ws_b = wb.create_sheet("Banco")
                cols_b_exp = ["DT_RECEB", "HR_RECEB", "TERMINAL", "CV",
                               "TXID", "VALOR", "saldo_rest", "status", "par_venda"]
                for ci, col in enumerate(cols_b_exp, 1):
                    c = ws_b.cell(1, ci, col.upper())
                    hdr_style(c, cor_hex="B45309")

                for ri, (_, row_d) in enumerate(self.df_banco.iterrows(), 2):
                    for ci, col in enumerate(cols_b_exp, 1):
                        v = row_d.get(col, "")
                        ws_b.cell(ri, ci, v)

                for ci, col in enumerate(cols_b_exp, 1):
                    ws_b.column_dimensions[get_column_letter(ci)].width = (
                        40 if col == "TXID" else 16)

            # ── Aba 4: Pendentes (itens sem conciliação) ──────────────────────
            ws_pend = wb.create_sheet("Pendentes")
            ws_pend["A1"] = "VENDAS PENDENTES"
            hdr_style(ws_pend["A1"])
            ws_pend.merge_cells("A1:G1")
            row_p = 2

            if self.df_vendas is not None:
                pend_v = self.df_vendas[self.df_vendas["status"] == "pendente"]
                for _, row_d in pend_v.iterrows():
                    for ci, col in enumerate(["origem","referencia","valor","saldo_rest","descricao","status","par_banco"], 1):
                        ws_pend.cell(row_p, ci, row_d.get(col, ""))
                    row_p += 1

            row_p += 1
            ws_pend.cell(row_p, 1, "PIX BANCO PENDENTES")
            hdr_style(ws_pend.cell(row_p, 1), cor_hex="B45309")
            ws_pend.merge_cells(f"A{row_p}:G{row_p}")
            row_p += 1

            if self.df_banco is not None:
                pend_b = self.df_banco[self.df_banco["status"] == "pendente"]
                for _, row_d in pend_b.iterrows():
                    for ci, col in enumerate(["DT_RECEB","HR_RECEB","TERMINAL","CV","TXID","VALOR","saldo_rest"], 1):
                        ws_pend.cell(row_p, ci, row_d.get(col, ""))
                    row_p += 1

            for ci in range(1, 8):
                ws_pend.column_dimensions[get_column_letter(ci)].width = 18
            ws_pend.column_dimensions["E"].width = 40

            wb.save(path_out)
            self.status_var.set(f"✅ Exportado: {os.path.basename(path_out)}")
            messagebox.showinfo("Exportação concluída",
                f"Relatório salvo com sucesso!\n\n{path_out}")

        except Exception as e:
            import traceback
            messagebox.showerror("Erro na exportação",
                f"{e}\n\n{traceback.format_exc()}")

    # ─── Helpers ─────────────────────────────────────────────────────────────

    def _item_id(self, tree, item):
        if tree == self.tree_v:
            return self._map_v.get(item)
        return self._map_b.get(item)

    def _buscar_item(self, tree, idx):
        if tree == self.tree_v:
            return self._rmap_v.get(idx)
        return self._rmap_b.get(idx)